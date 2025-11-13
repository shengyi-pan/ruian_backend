import warnings
from datetime import date, datetime, time
from decimal import Decimal
from typing import List, Optional

import pandas as pd
from openpyxl import load_workbook

from app.model.employee_worklog import EmployeeWorklog
from app.model.production_info import ProductionInfo

PRODUCTION_INFO_COLUMN_MAP = {
    "生产订单号": "order_no",
    "产品名称": "model",
    "单据编号": "brand_no",
    "单据日期": "doc_date",  # 中间字段，用于 created_at / updated_at
    "转出作业": "job_type",
    "合格数量": "quantity",
}

PRODUCTION_INFO_REQUIRED_COLUMNS = [
    "生产订单号",
    "产品名称",
    "单据编号",
    "单据日期",
    "转出作业",
    "合格数量",
]


def parse_production_excel(
    excel_path: str,
    sheet_name: Optional[str] = None,
    filter_month: Optional[str] = None,
) -> List[ProductionInfo]:
    """
    读取并解析 Excel, 返回 List[ProductionInfo]。

    参数：
        excel_path: Excel 文件路径
        sheet_name: 表单名称（默认读取第一个）
        filter_month: 月份过滤参数，格式为 yyyyMM (如 "202510")，默认为 None 表示不过滤

    逻辑：
      - 校验必要列是否存在
      - 行级清洗、类型转换
      - created_at/updated_at 来自「单据日期」
      - 对未映射字段使用默认值(如 performance_factor=1.00, upload_date=today, id=None)
      - 如果 filter_month 不为 None, 只保留 doc_date 与 filter_month 相同月份的数据
    """
    # 如果 sheet_name 为 None，默认读取第一个 sheet（索引 0）
    read_sheet = sheet_name if sheet_name is not None else 0
    # 抑制 openpyxl 的样式表警告
    with warnings.catch_warnings():
        warnings.filterwarnings("ignore", category=UserWarning, module="openpyxl")
        df = pd.read_excel(
            excel_path, sheet_name=read_sheet, dtype=str
        )  # 先以字符串读取，后续自行转换
    # 去除全空列名两侧空白
    df.columns = [str(c).strip() for c in df.columns]

    # 检查必要列
    missing = [c for c in PRODUCTION_INFO_REQUIRED_COLUMNS if c not in df.columns]
    if missing:
        raise ValueError(f"Excel 缺少必要列: {missing}")

    # 只保留关心的列
    keep_cols = [c for c in PRODUCTION_INFO_COLUMN_MAP.keys() if c in df.columns]
    work = df[keep_cols].copy()

    # 列重命名到中间/目标字段
    work = work.rename(columns=PRODUCTION_INFO_COLUMN_MAP)

    # 尽量把数量列转成数字（容错：有时 Excel 中是数字，读字符串时会有小数点）
    if "quantity" in work.columns:
        work["quantity"] = work["quantity"].apply(
            lambda x: None if pd.isna(x) else str(x).strip()
        )

    # 解析 filter_month（如果提供）
    filter_year = None
    filter_month_num = None
    if filter_month is not None:
        try:
            if len(filter_month) != 6:
                raise ValueError(
                    f"filter_month 格式错误，应为 yyyyMM 格式（6位），当前: {filter_month}"
                )
            filter_year = int(filter_month[:4])
            filter_month_num = int(filter_month[4:6])
            if not (1 <= filter_month_num <= 12):
                raise ValueError(
                    f"filter_month 月份无效，应为 01-12，当前: {filter_month[4:6]}"
                )
        except ValueError as e:
            raise ValueError(f"filter_month 参数格式错误: {e}")

    # 生成 Pydantic 模型
    results: List[ProductionInfo] = []
    for _, row in work.iterrows():
        doc_date_raw = row.get("doc_date", None)
        # 统一转换为 naive datetime，让 Pydantic validator 处理时区
        doc_datetime = _to_datetime(doc_date_raw)

        # 如果提供了 filter_month，进行月份过滤
        if filter_month is not None:
            if doc_datetime is None:
                # 无法解析日期，跳过该行
                continue

            # 比较年月
            if (
                doc_datetime.year != filter_year
                or doc_datetime.month != filter_month_num
            ):
                # 年月不匹配，跳过该行
                continue

        # 构建模型数据字典
        payload = {
            "order_no": row.get("order_no", ""),
            "model": row.get("model", ""),
            "brand_no": row.get("brand_no", ""),
            "quantity": row.get("quantity", None),
            "job_type": row.get("job_type", ""),
            # 未映射字段使用默认：performance_factor=1.00, upload_date=today, id=None
            "performance_factor": Decimal("1.00"),
            "upload_date": date.today(),
            "created_at": doc_datetime,  # 使用转换后的 naive datetime
            "updated_at": doc_datetime,  # 使用转换后的 naive datetime
        }

        # 创建并校验对象
        obj = ProductionInfo(**payload)
        results.append(obj)

    return results


WORKLOG_HEADER_TITLES = [
    "编号",
    "日期",
    "生产订单号",
    "数量",
    "圈数",
    "面系数",
    "绩效系数",
    "绩效数量",
]


def _to_order_no(value) -> str:
    """将订单号单元格统一转成字符串（去掉 .0 等）"""
    if value is None:
        return ""
    if isinstance(value, (int,)):
        return str(value)
    if isinstance(value, float):
        # 处理 2517281.0 这种情况
        if value.is_integer():
            return str(int(value))
        return str(value)
    return str(value).strip()


def _to_int(value) -> Optional[int]:
    if value is None or value == "":
        return None
    if isinstance(value, int):
        return value
    if isinstance(value, float):
        return int(value)
    try:
        return int(str(value).strip())
    except ValueError:
        return None


def _to_float(value) -> Optional[float]:
    if value is None or value == "":
        return None
    if isinstance(value, (int, float)):
        return float(value)
    try:
        return float(str(value).strip())
    except ValueError:
        return None


def _to_datetime(value) -> Optional[datetime]:
    if value is None or value == "":
        return None

    # 处理 pandas Timestamp，转换为 naive datetime
    if isinstance(value, pd.Timestamp):
        # 转换为 Python datetime，如果是 timezone-aware 则转换为本地时间后移除时区信息
        dt = value.to_pydatetime()
        if dt.tzinfo is not None:
            # 移除时区信息，保留时间值
            dt = dt.replace(tzinfo=None)
        return dt

    if isinstance(value, datetime):
        # 如果是 timezone-aware datetime，移除时区信息
        if value.tzinfo is not None:
            return value.replace(tzinfo=None)
        return value
    if isinstance(value, date):
        return datetime.combine(value, time.min)

    # 字符串日期，如 "2025/10/1" 或 "2025-10-01"
    text = str(value).strip()
    for fmt in ("%Y/%m/%d", "%Y-%m-%d", "%Y.%m.%d"):
        try:
            d = datetime.strptime(text, fmt)
            return d
        except ValueError:
            continue
    # 实在解析不了就返回 None
    return None


def parse_sheet(ws, employee_id: str) -> List[EmployeeWorklog]:
    """
    解析单个 sheet, 返回该员工的所有工作量记录
    """
    worklogs: List[EmployeeWorklog] = []

    header_col_index = None  # {列名: index}

    for row in ws.iter_rows(values_only=True):
        # 检测表头行
        first_cell = row[0]
        if isinstance(first_cell, str) and first_cell.strip() == "编号":
            # 动态记录各列位置，防止列顺序调整
            header_col_index = {}
            for idx, cell in enumerate(row):
                if isinstance(cell, str):
                    title = cell.strip()
                    if title in WORKLOG_HEADER_TITLES:
                        header_col_index[title] = idx
            # 找到表头后，继续下一行
            continue

        # 没遇到任何表头之前的行全部跳过
        if not header_col_index:
            continue

        # 空行跳过
        if all(c is None or str(c).strip() == "" for c in row):
            continue

        # 合计行 / 汇总行跳过（一般订单号为空）
        # 另外也可以根据第一个非空单元格是“合计”“汇总”来判断
        # 这里简单判断订单号是否为空：为空就跳过
        order_no_idx = header_col_index.get("生产订单号")
        order_cell = row[order_no_idx] if order_no_idx is not None else None
        order_no = _to_order_no(order_cell)
        if not order_no:
            # 例如合计行、说明行等
            continue

        # 取其它列
        date_idx = header_col_index.get("日期")
        qty_idx = header_col_index.get("数量")
        perf_factor_idx = header_col_index.get("绩效系数")
        perf_amount_idx = header_col_index.get("绩效数量")

        date_cell = row[date_idx] if date_idx is not None else None
        qty_cell = row[qty_idx] if qty_idx is not None else None
        perf_factor_cell = row[perf_factor_idx] if perf_factor_idx is not None else None
        perf_amount_cell = row[perf_amount_idx] if perf_amount_idx is not None else None

        work_date = _to_datetime(date_cell)
        if work_date is None:
            # 没有日期就跳过该行，避免脏数据
            continue

        quantity = _to_int(qty_cell)
        if quantity is None or quantity <= 0:
            quantity = 1

        perf_factor = _to_float(perf_factor_cell)
        if perf_factor is None or perf_factor <= 0:
            # 若缺失则给默认 1.0（满足 >0 的约束）
            perf_factor = 1.0

        perf_amount = _to_float(perf_amount_cell)
        if perf_amount is None or perf_amount <= 0:
            # 若绩效数量没填，则先跳过该行，避免脏数据
            continue

        # 构造 BaseModel 实例
        worklog = EmployeeWorklog(
            order_no=order_no,
            model=None,
            brand_no=None,
            employee_id=str(employee_id).strip(),
            employee_name=None,
            job_type="未知",
            quantity=quantity,
            performance_factor=perf_factor,
            performance_amount=perf_amount,
            work_date=work_date,
        )
        worklogs.append(worklog)

    return worklogs


def parse_employee_worklogs_from_excel(file_path: str) -> List[EmployeeWorklog]:
    """
    解析整个 Excel, 返回所有员工的工作量数据 (List[EmployeeWorklog])
    每个 sheet 的名字就是 employee_id
    """
    # 抑制 openpyxl 的样式表警告
    with warnings.catch_warnings():
        warnings.filterwarnings("ignore", category=UserWarning, module="openpyxl")
        wb = load_workbook(filename=file_path, data_only=True)
    all_worklogs: List[EmployeeWorklog] = []

    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        sheet_logs = parse_sheet(ws, employee_id=sheet_name)
        all_worklogs.extend(sheet_logs)

    return all_worklogs


if __name__ == "__main__":
    # # 你的文件路径（示例中为你上传的文件名；若脚本与文件同目录可直接使用）
    # excel_file = "/Users/sy.pan/Documents/workspace/ml_lesson/ruian_backend/data/product_info_demo.xlsx"

    # items = parse_production_excel(excel_file, filter_month="202511")
    # print(f"共解析到 {len(items)} 条记录")
    # # 打印前3条示例
    # for i, it in enumerate(items[:3], 1):
    #     print(f"[{i}] -> {it.model_dump()}")

    # 示例：解析当前目录下的 10月毛刷台账.xlsx
    file_path = "/Users/sy.pan/Documents/workspace/ml_lesson/ruian_backend/data/worklog_demo.xlsx"
    logs = parse_employee_worklogs_from_excel(file_path)
    print(f"总共解析出 {len(logs)} 条记录")
    # 打印前几条示意
    for item in logs[:5]:
        print(item.model_dump())
